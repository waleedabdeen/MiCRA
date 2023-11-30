import glob
import xlrd
import re
import csv
from openpyxl import load_workbook
import unicodedata as ud
from treelib import Node, Tree

ROOT_NAME = "root"
BASE_PATH = "../data"

def coclass_from_df(df:object, name: str, content_index: int):
    tree = Tree()
    tree.create_node(name, ROOT_NAME)
    previousTable = ''
    for index, row in df.iterrows():
        table_name = row.iloc[0].strip()
        identifier = row.iloc[1].strip()
        content = row.iloc[content_index].strip()

        if len(identifier) == 1 and previousTable != table_name: 
            previousTable = table_name
            tree.create_node(table_name, table_name, ROOT_NAME)
        if len(identifier) == 1:
            tree.create_node(content, table_name + identifier, table_name)
        else:
            tree.create_node(content, table_name + identifier, table_name + identifier[:-1])
    
    return tree

def sb11_from_df(df:object, name: str, content_index:int):
    # Needed to modify the original file since three tables were stored in a single sheet and the items
    # of the second and third table were not ordered.

    # Fixed typos:
    # alternativtabell: U261 -> U-261
    tree = Tree()
    tree.create_node(name, ROOT_NAME)

    previousTable = ''
    for index, row in df.iterrows():
        table_name = str(row.iloc[0]).strip()
        identifier = str(row.iloc[1]).strip()
        content = row.iloc[content_index].strip()

        if previousTable != table_name: 
            previousTable = table_name
            tree.create_node(table_name, table_name, ROOT_NAME)

        if table_name == "Byggdelar":
            if len(identifier) == 1:
                parent = table_name
            else:
                parent = table_name + identifier[:-1]
            tree.create_node(content, table_name + identifier, parent)
        elif table_name == "Alternativtabell":
            if identifier[0] == "U": # Parts of the table are hierarchical (U-xxx), the rests isn't
                identifier = identifier.replace("-", "")

                if identifier == "U":
                    tree.create_node(content, table_name + identifier, table_name)
                else:
                    tree.create_node(content, table_name + identifier, table_name + identifier[:-1])
            else: #The non-hierarchical items
                tree.create_node(content, table_name + identifier, table_name)
        elif table_name == "Landskapsinformation":
            #These are all flat items, no hierarchy
            tree.create_node(content, table_name + identifier, table_name)
            
    return tree

def coclass(name):
    tree = Tree();
    tree.create_node(name, ROOT_NAME)

    # Using the cvs file since I have already added there dummy nodes that are missing
    # in the original for a proper tree:
    # Tillgångssystem: Dummy*
    # Added also dummy IDs: 1, 2, 3
    with open(f'{BASE_PATH}/coclass_en_sv.csv') as csvfile:
        reader = csv.reader(csvfile, delimiter=';')

        #Skip header
        next(reader)

        previousTable = ''
        for row in reader:
            table_name = row[0].strip()
            identifier = row[1].strip()
            content = row[2].strip()

            if len(identifier) == 1 and previousTable != table_name: 
                previousTable = table_name
                tree.create_node(table_name, table_name, ROOT_NAME)
            if len(identifier) == 1:
                tree.create_node(content, table_name + identifier, table_name)
            else:
                tree.create_node(content, table_name + identifier, table_name + identifier[:-1])

    return tree

def sb11(name):
    # Needed to modify the original file since three tables were stored in a single sheet and the items
    # of the second and third table were not ordered.

    # Fixed typos:
    # alternativtabell: U261 -> U-261

    path = f'{BASE_PATH}/sb11'

    tree = Tree()
    tree.create_node(name, ROOT_NAME)

    wb = load_workbook(f'{path}/SB11 CAD-Lager_ Elementkod_2020-11-27 13_46_19.xlsx')
    for ws_name in wb.sheetnames:
        if ws_name == "Original":
            continue

        tree.create_node(ws_name, ws_name, ROOT_NAME)
        ws = wb[ws_name]
        data = ws.values

        #skip header
        next(data)

        for row in data:
            identifier = str(row[0])
            content = row[1].strip()
            table_name = ws_name

            if table_name == "Byggdelar":
                if len(identifier) == 1:
                    parent = table_name
                else:
                    parent = identifier[:-1]
                tree.create_node(content, identifier, parent)
            elif table_name == "Alternativtabell":
                if identifier[0] == "U": # Parts of the table are hierarchical (U-xxx), the rests isn't
                    identifier = identifier.replace("-", "")

                    if identifier == "U":
                        tree.create_node(content, table_name + identifier, table_name)
                    else:
                        tree.create_node(content, table_name + identifier, table_name + identifier[:-1])
                else: #The non-hierarchical items
                    tree.create_node(content, table_name + identifier, table_name)
            elif table_name == "Landskapsinformation":
                #These are all flat items, no hierarchy
                tree.create_node(content, table_name + identifier, table_name)
    return tree

def path_to_root(tree, node, path):
    path.append(node)
    if not node.is_root():
        path_to_root(tree, tree.parent(node.identifier), path)
    return path

def depth(cs, dimension):
    SB11_depth = { 'Landskapsinformation': 1, 'Alternativtabell': 5, 'Byggdelar': 6 }
    CoClass_depth = {'Tillgångssystem' : 3, 'Konstruktiva system': 3, 'Grundfunktioner och Komponenter': 4 }

    if cs == 'SB11':
        return SB11_depth[dimension]
    elif cs == 'CoClass':
        return CoClass_depth[dimension]
    else:
        raise Exception('invalid cs name')
